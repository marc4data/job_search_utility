"""compile_profile.py — compile the Profile Workbook into the engine's files.

The Profile Workbook (one friendly .xlsx the user edits) is the single source of
truth. This compiles it into the three files the engine already consumes —
profile.py, bases.py, verified_skills.md — in their EXACT existing formats, so the
engine itself is unchanged. Those files are generated artifacts; edit the workbook.

Tabs consumed:
  Identity        -> profile.py
  Skills Matrix   -> verified_skills.md (tool truth) + the validator
  Experience      -> verified_skills.md (per-company domain/work/leadership)
  Career Highlights, Roles, Role Bullets, Tech Expertise, Projects, Bases Config
                  -> bases.py  (LEADER + HANDSON)

Usage:  python3 compile_profile.py <workbook.xlsx> <out_profile_dir>
"""
import os
import sys
import openpyxl


def rows(ws):
    """Read a sheet as a list of dicts keyed by the row-1 header."""
    headers = [(c.value or "").strip() if isinstance(c.value, str) else c.value for c in ws[1]]
    out = []
    for r in range(2, ws.max_row + 1):
        vals = {}
        empty = True
        for i, h in enumerate(headers, start=1):
            if not h:
                continue
            v = ws.cell(row=r, column=i).value
            v = "" if v is None else (str(v).strip() if isinstance(v, str) else v)
            vals[h] = v
            if v != "":
                empty = False
        if not empty:
            out.append(vals)
    return out


def _pylist(items, indent=8):
    pad = " " * indent
    return "\n".join(f"{pad}{i!r}," for i in items)


def compile_profile_py(wb, out_dir):
    ident = {r["Field"]: r["Value"] for r in rows(wb["Identity"])}
    links = [(k.replace("Link: ", ""), v) for k, v in ident.items() if k.startswith("Link: ")]
    edu = [v for k, v in ident.items() if k.startswith("Education") and v]
    contact_bits = [ident.get("Location", ""), ident.get("Phone", ""), ident.get("Email", "")]
    contact = "  •  ".join(b for b in contact_bits if b)
    cl_contact = "  |  ".join(b for b in contact_bits if b)
    certs = []
    if ident.get("Certifications"):
        certs.append(f"Certifications: {ident['Certifications']}")
    if ident.get("Training"):
        certs.append(f"Training: {ident['Training']}")
    addl = [v for k, v in ident.items() if k.startswith("Additional Experience") and v]

    links_py = "\n".join(f'        ("{n}", "{u}"),' for n, u in links)
    text = f'''# profile.py — GENERATED from the Profile Workbook. Do not hand-edit; edit the workbook.
PROFILE = {{
    "name": {ident.get("Name","")!r},
    "contact": {contact!r},
    "cl_contact": {cl_contact!r},
    "links": [
{links_py}
    ],
    "education": [
{_pylist(edu)}
    ],
    "certifications": [
{_pylist(certs)}
    ],
    "additional_experience": [
{_pylist(addl)}
    ],
}}
'''
    open(os.path.join(out_dir, "profile.py"), "w").write(text)


def _bullets_for(bullet_rows, base, company):
    return [r["Bullet"] for r in bullet_rows
            if r.get("Base") == base and r.get("Company • Loc") == company and r.get("Bullet")]


def compile_bases_py(wb, out_dir):
    highlights = rows(wb["Career Highlights"])
    roles = rows(wb["Roles"])
    bullets = rows(wb["Role Bullets"])
    techs = rows(wb["Tech Expertise"])
    projects = rows(wb["Projects"]) if "Projects" in wb.sheetnames else []
    config = {r["Base"]: r for r in rows(wb["Bases Config"])}

    def base_block(base):
        ch = [r["Highlight"] for r in highlights if r.get("Base") == base and r.get("Highlight")]
        exp = []
        for r in roles:
            if r.get("Base") != base:
                continue
            company = r.get("Company • Loc", "")
            exp.append({"role": r.get("Role", ""), "company_loc": company,
                        "dates": r.get("Dates", ""), "desc": r.get("Desc", ""),
                        "bullets": _bullets_for(bullets, base, company)})
        te = [(r["Label"], r["Items"]) for r in techs if r.get("Base") == base and r.get("Label")]
        # projects grouped by name, in row order
        projs, seen = [], {}
        for r in projects:
            if r.get("Base") and r["Base"] != base:
                continue
            name = r.get("Project", "")
            if not name:
                continue
            if name not in seen:
                seen[name] = {"name": name, "tech": r.get("Tech", ""), "bullets": []}
                projs.append(seen[name])
            if r.get("Bullet"):
                seen[name]["bullets"].append(r["Bullet"])
        certs_before = str(config.get(base, {}).get("certs_before_edu", "")).lower() == "yes"
        return ch, exp, te, (projs or None), certs_before

    def fmt_block(base):
        ch, exp, te, projs, certs_before = base_block(base)
        exp_py = ",\n".join(
            "            {" + f'"role": {e["role"]!r}, "company_loc": {e["company_loc"]!r},\n'
            f'             "dates": {e["dates"]!r}, "desc": {e["desc"]!r},\n'
            f'             "bullets": [\n' + "".join(f"                 {b!r},\n" for b in e["bullets"]) +
            "             ]}" for e in exp)
        te_py = "\n".join(f"            ({l!r}, {i!r})," for l, i in te)
        if projs:
            proj_py = "[\n" + ",\n".join(
                "            {" + f'"name": {p["name"]!r},\n             "tech": {p["tech"]!r},\n'
                f'             "bullets": [\n' + "".join(f"                 {b!r},\n" for b in p["bullets"]) +
                "             ]}" for p in projs) + "\n        ]"
        else:
            proj_py = "None"
        return (f'    "{base}": {{\n'
                f'        "career_highlights": [\n' + "".join(f"            {h!r},\n" for h in ch) + "        ],\n"
                f'        "experience": [\n{exp_py}\n        ],\n'
                f'        "tech_expertise": [\n{te_py}\n        ],\n'
                f'        "projects": {proj_py},\n'
                f'        "certs_before_edu": {certs_before},\n'
                f'    }},')

    bases_order = [b for b in ("LEADER", "HANDSON") if b in {r["Base"] for r in roles}]
    body = "\n".join(fmt_block(b) for b in bases_order)
    text = ("# bases.py — GENERATED from the Profile Workbook. Do not hand-edit; edit the workbook.\n"
            "BASES = {\n" + body + "\n}\n")
    open(os.path.join(out_dir, "bases.py"), "w").write(text)


def compile_verified_skills(wb, out_dir):
    exp = {r["Company"]: r for r in rows(wb["Experience"])}
    per_company, portfolio, confirm, never = {}, [], [], []
    for r in rows(wb["Skills Matrix"]):
        skill, company = r.get("Skill / Tool", ""), r.get("Company", "")
        if not skill:
            continue
        if str(r.get("Never-used?", "")).lower() == "yes":
            never.append(skill)
        elif company:
            per_company.setdefault(company, []).append(skill)
        elif str(r.get("Portfolio-only?", "")).lower() == "yes":
            portfolio.append((skill, r.get("Notes", "")))
        else:
            confirm.append((skill, r.get("Notes", "")))

    out = ["# Verified Skills — GENERATED from the Profile Workbook. Edit the workbook, not this file.",
           "", "Never claim a tool unless it appears below for the relevant company.", "", "---", "",
           "## Tools / skills NEVER used (never claim these)"]
    out += [f"- {t}" for t in never] or ["- _(none)_"]
    out += ["", "## Skills that are PORTFOLIO / LEARNING ONLY (claim only with that attribution)"]
    out += [f"- **{t}** — {n or 'portfolio project only'}" for t, n in portfolio] or ["- _(none)_"]
    out += ["", "## Confirm employer attribution (not yet tied to a company)"]
    out += [f"- {t}{(' — ' + n) if n else ''}" for t, n in confirm] or ["- _(none)_"]
    out += ["", "---", "", "## Per-company verified experience", ""]
    for company, tools in per_company.items():
        m = exp.get(company, {})
        head = f"### {company}" + (f"  ({m.get('Title','')} — {m.get('Dates','')})" if m else "")
        out += [head, f"- **Stack / tools:** {', '.join(tools)}",
                f"- **Domain:** {m.get('Domain','')}",
                f"- **Type of work:** {m.get('What you did','')}",
                f"- **Leadership:** {m.get('Leadership','')}", ""]
    open(os.path.join(out_dir, "verified_skills.md"), "w").write("\n".join(out) + "\n")


def compile_all(workbook_path, out_dir):
    os.makedirs(out_dir, exist_ok=True)
    wb = openpyxl.load_workbook(workbook_path)
    compile_profile_py(wb, out_dir)
    compile_bases_py(wb, out_dir)
    compile_verified_skills(wb, out_dir)


if __name__ == "__main__":
    if len(sys.argv) < 3:
        sys.exit("usage: python3 compile_profile.py <workbook.xlsx> <out_profile_dir>")
    compile_all(sys.argv[1], sys.argv[2])
    print(f"Compiled {sys.argv[1]} → {sys.argv[2]}/(profile.py, bases.py, verified_skills.md)")
