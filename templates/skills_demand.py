"""skills_demand.py — the skills-demand repository (Epic R / v2).

A persistent, accumulating record of what employers ask for across every processed
job, categorized (tool / technical_competency / leadership / domain), normalized,
and surfaced as a Covered/Weak/Gap review against the user's profile.

  record_jd()             persist one JD (readable filename) + its categorized skills (R1/R2/R3)
  build_index_report()    demand inventory by category (R2/R4)
  write_demand_workbook() the Excel "Skills Demand" review tab (R5)

Demand data only — never writes the Skills Matrix, bases.py, or verified_skills.md.
The CSV joins on a stable job_id, so re-processing updates rather than duplicates.
"""
import csv
import datetime
import os
import re
import sys

from skills_taxonomy import extract, NONTOOL, display
from validate_profile import load_truth

FIELDS = ["job_id", "date", "company", "role", "level", "category", "skill", "raw"]
NO_EXTRACT = "(no_extract)"


def _corpus_dir(home):
    d = os.path.join(home, "docs", "job_descriptions")
    os.makedirs(d, exist_ok=True)
    return d


def _index_path(home):
    return os.path.join(_corpus_dir(home), "skills_demand_index.csv")


def _slug(s, cap=60):
    s = re.sub(r"[^a-z0-9\s-]", "", (s or "").lower())
    s = re.sub(r"\s+", "-", s.strip())
    return re.sub(r"-+", "-", s)[:cap].strip("-") or "untitled"


def _corpus_filename(home, date, company, role, job_id):
    """R1: YYYY-MM-DD_company_jobtitle.md; re-processing the same job_id updates
    the same file; a true collision (different job_id) appends job_id."""
    base = f"{date}_{_slug(company)}_{_slug(role)}"
    plain = os.path.join(_corpus_dir(home), base + ".md")
    if os.path.exists(plain):
        first = open(plain).readline()
        owner = re.search(r"job_id:\s*(\S+)", first)
        if owner and owner.group(1) != job_id:        # real collision
            return os.path.join(_corpus_dir(home), f"{base}__{job_id}.md")
    return plain


def load_index(home):
    path = _index_path(home)
    if not os.path.exists(path):
        return []
    with open(path, newline="") as f:
        return list(csv.DictReader(f))


def _write_index(home, rows):
    with open(_index_path(home), "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=FIELDS)
        w.writeheader(); w.writerows(rows)


def record_jd(home, job_id, company, role, level, jd_text, date=None):
    """Persist one retrieved JD + its categorized skills. Returns the rows written.

    R3: a processed role ALWAYS contributes >=1 row — if nothing extracts, a single
    explicit no_extract marker (so the gap is visible, never a silent drop)."""
    date = date or datetime.date.today().isoformat()

    # Remove any old corpus file for this job whose name changed, then write fresh.
    for fn in os.listdir(_corpus_dir(home)):
        if fn.endswith(".md"):
            p = os.path.join(_corpus_dir(home), fn)
            if re.search(rf"job_id:\s*{re.escape(job_id)}\b", open(p).readline()):
                os.remove(p)
    path = _corpus_filename(home, date, company, role, job_id)
    with open(path, "w") as f:
        f.write(f"<!-- job_id: {job_id} -->\n# {company} — {role} ({level}) · {date}\n\n{jd_text}\n")

    extracted = extract(jd_text)
    base = {"job_id": job_id, "date": date, "company": company, "role": role, "level": level}
    if extracted:
        rows_new = [{**base, "category": cat, "skill": canon, "raw": raw}
                    for cat, canon, raw in extracted]
    else:
        rows_new = [{**base, "category": "none", "skill": NO_EXTRACT, "raw": ""}]   # R3 marker

    rows = [r for r in load_index(home) if r["job_id"] != job_id] + rows_new   # R1: update, not dup
    _write_index(home, rows)
    return rows_new


def demand(rows):
    """skill -> {category, jobs:set, levels:set}, excluding no_extract markers."""
    agg = {}
    for r in rows:
        if r["skill"] == NO_EXTRACT:
            continue
        key = (r["category"], r["skill"])
        d = agg.setdefault(key, {"jobs": set(), "levels": set()})
        d["jobs"].add(r["job_id"]); d["levels"].add(r["level"])
    return agg


def reconcile(rows):
    """R3 auditability: processed job count vs jobs that produced a real skill."""
    all_jobs = {r["job_id"] for r in rows}
    no_extract_jobs = {r["job_id"] for r in rows if r["skill"] == NO_EXTRACT}
    return {"processed": len(all_jobs), "no_extract": len(no_extract_jobs)}


# ── coverage of a demanded skill against the user's profile ──────────────────
def _nontool_evidence(profile_dir):
    """Lowercased blob of verified_skills.md + bases.py for non-tool coverage checks."""
    blob = ""
    for fn in ("verified_skills.md", "bases.py"):
        p = os.path.join(profile_dir, fn)
        if os.path.exists(p):
            blob += "\n" + open(p).read().lower()
    return blob


def coverage(category, skill, truth, evidence):
    backed, portfolio, confirm, never = truth
    if category == "tool":
        if skill in backed:
            return "Covered"
        if skill in portfolio or skill in confirm:
            return "Weak"
        return "Gap"
    # non-tool: Covered if any synonym phrase appears in the user's own profile text
    phrases = NONTOOL.get(category, {}).get(skill, [skill])
    return "Covered" if any(ph in evidence for ph in phrases) else "Gap"


def build_index_report(home, profile_dir):
    rows = load_index(home)
    if not rows:
        return "No JDs recorded yet — process some opportunities first."
    agg = demand(rows)
    truth = load_truth(profile_dir)
    evidence = _nontool_evidence(profile_dir)
    rec = reconcile(rows)
    n = rec["processed"]
    out = [f"SKILLS-DEMAND REPOSITORY · {n} jobs processed "
           f"({rec['no_extract']} produced no extractable skill) · {_index_path(home)}",
           "=" * 72]
    for cat in ("tool", "technical_competency", "leadership", "domain"):
        items = sorted(((s, d) for (c, s), d in agg.items() if c == cat),
                       key=lambda kv: -len(kv[1]["jobs"]))
        if not items:
            continue
        out.append(f"\n[{cat}]  (skill — jobs — coverage)")
        for skill, d in items:
            cov = coverage(cat, skill, truth, evidence)
            out.append(f"  {display(cat, skill):<26} {len(d['jobs'])}/{n:<3} {cov}")
    return "\n".join(out)


def write_demand_workbook(home, profile_dir):
    """R5: an Excel 'Skills Demand' tab — frequency + %jobs + Covered/Weak/Gap,
    grouped by category, non-tool categories included. Read-only output."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    rows = load_index(home)
    agg = demand(rows)
    truth = load_truth(profile_dir)
    evidence = _nontool_evidence(profile_dir)
    n = reconcile(rows)["processed"] or 1

    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Skills Demand"
    headers = ["Category", "Skill", "Jobs", "% of jobs", "Coverage", "Levels"]
    widths = [22, 28, 8, 10, 12, 22]
    fills = {"Covered": "FFE2EFDA", "Weak": "FFFFF2CC", "Gap": "FFFCE4D6"}
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFFFF"); c.fill = PatternFill("solid", fgColor="FF1F3864")
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    r = 2
    for cat in ("tool", "technical_competency", "leadership", "domain"):
        items = sorted(((s, d) for (c, s), d in agg.items() if c == cat),
                       key=lambda kv: -len(kv[1]["jobs"]))
        for skill, d in items:
            cov = coverage(cat, skill, truth, evidence)
            vals = [cat, display(cat, skill), len(d["jobs"]),
                    round(100 * len(d["jobs"]) / n), cov, ", ".join(sorted(d["levels"]))]
            for col, v in enumerate(vals, start=1):
                cell = ws.cell(row=r, column=col, value=v); cell.alignment = Alignment(vertical="top")
            ws.cell(row=r, column=5).fill = PatternFill("solid", fgColor=fills[cov])
            r += 1
    out = os.path.join(_corpus_dir(home), "skills_demand.xlsx")
    wb.save(out)
    return out


if __name__ == "__main__":
    home = sys.argv[1] if len(sys.argv) > 1 else "."
    profile_dir = sys.argv[2] if len(sys.argv) > 2 else os.path.join(home, "profile")
    print(build_index_report(home, profile_dir))
    print("\nSkills Demand workbook →", write_demand_workbook(home, profile_dir))
