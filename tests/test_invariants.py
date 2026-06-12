"""Regression tests for the job-search-tailor invariants (backlog #4).

(a) critique_and_refine() never writes a tracker.
(b) A dummy-profile build renders the profile name and leaves no placeholder text.
(c) The ats_score scorer writes the expected "Resume Score" column.
"""
import glob
import os

import openpyxl
import pytest

from conftest import DUMMY_NAME_UPPER


def _sample_resume_args():
    """Minimal, valid arguments for build_resume / critique_and_refine."""
    summary_lines = [
        "Analytics leader with deep python and sql experience.",
        "Builds executive dashboards and reliable data pipelines.",
    ]
    expertise_items = ["Python", "SQL", "Dashboards", "Data Quality"]
    return summary_lines, expertise_items


# ── (a) critique_and_refine never writes a tracker ──────────────────────────
def test_critique_and_refine_writes_no_tracker(engine, tmp_path):
    summary_lines, expertise_items = _sample_resume_args()
    resume_path = str(tmp_path / "resume.docx")
    engine.build_resume(resume_path, "Director of Analytics", summary_lines, expertise_items)

    # A sentinel .xlsx in the working dir: if any code path naively wrote a
    # tracker, it would most likely clobber or create one here.
    sentinel = tmp_path / "job_tracker.xlsx"
    wb = openpyxl.Workbook()
    wb.active["A1"] = "do-not-touch"
    wb.save(str(sentinel))
    before_bytes = sentinel.read_bytes()
    xlsx_before = set(glob.glob(str(tmp_path / "**" / "*.xlsx"), recursive=True))

    role_dict = {
        "jd_hard": [("python", 3), ("sql", 2), ("databricks", 3)],  # one real miss
        "gap_patches": {},   # nothing addressable → forces the no-patch path
        "tight_patches": {},
        "company": "Test Co",
        "role": "Director of Analytics",
    }
    score = engine.critique_and_refine(
        resume_path, "Test Co", "Director of Analytics", "Director of Analytics",
        summary_lines, expertise_items, role_dict,
    )

    assert isinstance(score, int)                          # it returns a score...
    assert sentinel.read_bytes() == before_bytes           # ...but touches no tracker
    xlsx_after = set(glob.glob(str(tmp_path / "**" / "*.xlsx"), recursive=True))
    assert xlsx_after == xlsx_before                       # and creates no new .xlsx


def test_engine_exposes_no_tracker_write_helpers(engine):
    """The removed foot-gun functions must stay gone."""
    for name in ("score_and_log", "_write_tracker_score", "ats_score_resume"):
        assert not hasattr(engine, name), f"{name} reintroduced into the engine"


# ── (b) dummy-profile build renders the name and no placeholders ────────────
def test_build_renders_name_without_placeholders(engine, tmp_path, docx_text):
    summary_lines, expertise_items = _sample_resume_args()
    resume_path = str(tmp_path / "resume.docx")
    engine.build_resume(resume_path, "Director of Analytics", summary_lines, expertise_items)

    assert os.path.exists(resume_path)
    text = docx_text(resume_path)

    # The profile name is rendered (header upper-cases it).
    assert DUMMY_NAME_UPPER in text

    # No template/placeholder residue leaked into the document.
    lowered = text.lower()
    for token in ("{{", "}}", "todo", "placeholder", "lorem ipsum",
                  "your name", "your university", "example.com"):
        assert token not in lowered, f"placeholder token {token!r} found in resume"


# ── (c) the ats_score scorer writes the expected column ─────────────────────
def _make_tracker(path, headers, sheet_name="Applications"):
    """Build a tracker .xlsx with the header row on row 3 (the project layout)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    for col, head in enumerate(headers, start=1):
        ws.cell(row=3, column=col, value=head)
    ws.cell(row=4, column=1, value="Example Co")   # a seeded data row
    wb.save(str(path))


def test_ats_scorer_writes_resume_score_column(ats_module, tmp_path, monkeypatch):
    tracker = tmp_path / "tracker.xlsx"
    headers = ["Company", "Role / Title", "Resume Score", "Status"]
    _make_tracker(tracker, headers, sheet_name="Applications")

    # Point the template's module-level constants at our temp tracker.
    monkeypatch.setattr(ats_module, "TRACKER", str(tracker))
    monkeypatch.setattr(ats_module, "SHEET", "Applications")

    ats_module.write_tracker(4, 87)

    wb = openpyxl.load_workbook(str(tracker))
    ws = wb["Applications"]
    score_col = [c.value for c in ws[3]].index("Resume Score") + 1
    assert ws.cell(row=4, column=score_col).value == 87
    # Nothing else on the row was disturbed.
    assert ws.cell(row=4, column=1).value == "Example Co"


def test_ats_scorer_fails_clearly_without_score_column(ats_module, tmp_path, monkeypatch):
    tracker = tmp_path / "tracker_noscore.xlsx"
    _make_tracker(tracker, ["Company", "Role / Title", "Status"], sheet_name="Applications")
    monkeypatch.setattr(ats_module, "TRACKER", str(tracker))
    monkeypatch.setattr(ats_module, "SHEET", "Applications")

    with pytest.raises(SystemExit):
        ats_module.write_tracker(4, 87)


def test_ats_score_is_weighted_and_honest(ats_module):
    """A partial keyword match yields a partial (not inflated) score."""
    resume_text = "experienced in python and sql for analytics"
    jd_terms = [("python", 3), ("sql", 2), ("databricks", 3), ("kafka", 2)]
    score, found, missed = ats_module.ats_score(resume_text, jd_terms)
    # hit 3+2 of 3+2+3+2 = 5/10 → 50
    assert score == 50
    assert "databricks[3]" in missed and "kafka[2]" in missed
