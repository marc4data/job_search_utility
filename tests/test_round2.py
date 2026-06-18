"""Regression tests for Round 2 (v0.2.0): versioning banner (K1), JD-retrieval
helper (G1/G2/G3), and the mandatory batch summary table (J1).

Prompt-only behavior (the model actually fetching a JD, the live paste UX) is
validated in clean-room runs, not here. These tests lock the deterministic,
testable surface: the version source of truth, the retrieval helper's pure
functions, and the table the ATS script emits.
"""
import json
import os
import re

from conftest import _load_module

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
PLUGIN_JSON = os.path.join(REPO_ROOT, ".claude-plugin", "plugin.json")
SKILLS = [
    os.path.join(REPO_ROOT, "skills", "process-opportunities", "SKILL.md"),
    os.path.join(REPO_ROOT, "skills", "setup-profile", "SKILL.md"),
]

# A semver like 1.2.3 or a v-prefixed one — what a *hardcoded* banner would contain.
_VERSION_LITERAL = re.compile(r"\bv?\d+\.\d+\.\d+\b")


# ── K1: versioning + runtime banner ─────────────────────────────────────────
def test_plugin_version_is_semver_and_in_changelog():
    with open(PLUGIN_JSON) as f:
        version = json.load(f)["version"]
    assert re.fullmatch(r"\d+\.\d+\.\d+", version), version
    # the running version must have a CHANGELOG entry (single source of truth)
    changelog = open(os.path.join(REPO_ROOT, "CHANGELOG.md")).read()
    assert f"## [{version}]" in changelog, f"CHANGELOG.md missing an entry for {version}"


def test_skills_read_version_from_plugin_json_not_hardcoded():
    for path in SKILLS:
        with open(path) as f:
            text = f.read()
        # Each skill must point at the single source of truth...
        assert "plugin.json" in text, f"{path} does not reference plugin.json"
        # ...and must NOT bake in a concrete version that would drift.
        leaked = _VERSION_LITERAL.findall(text)
        assert not leaked, f"{path} contains a hardcoded version literal: {leaked}"


# ── G1/G2/G3: the JD-retrieval helper ───────────────────────────────────────
def _jd():
    return _load_module(
        "jd_retrieval_under_test",
        os.path.join(REPO_ROOT, "templates", "jd_retrieval.py"),
    )


# G1 — read the embedded hyperlink target, not the display text
def test_resolve_cell_link_prefers_hyperlink_target(tmp_path):
    import openpyxl

    jd = _jd()
    wb = openpyxl.Workbook()
    ws = wb.active
    # A cell whose *display* text is "Linkedin" but whose hyperlink is the URL.
    c = ws.cell(row=4, column=1, value="Linkedin")
    c.hyperlink = "https://www.linkedin.com/jobs/view/4012345678"
    assert jd.resolve_cell_link(c) == "https://www.linkedin.com/jobs/view/4012345678"

    # No hyperlink, but a URL sitting in the text → fall back to it.
    c2 = ws.cell(row=5, column=1, value="see https://acme.example/careers/123 today")
    assert jd.resolve_cell_link(c2) == "https://acme.example/careers/123"

    # Pure display text, no link anywhere → nothing to fetch.
    c3 = ws.cell(row=6, column=1, value="DM Message")
    assert jd.resolve_cell_link(c3) is None


def test_find_link_column_by_header():
    jd = _jd()
    headers = ["#", "Company", "Resume Score", "Sourced From (w/link)", "Status"]
    assert jd.find_link_column(headers) == 4          # exact header
    assert jd.find_link_column(["a", "Job Link", "b"]) == 2  # contains 'link'
    assert jd.find_link_column(["a", "b", "c"]) is None      # absent → None


def test_read_job_links_end_to_end(tmp_path):
    import openpyxl

    jd = _jd()
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = ["Company", "Sourced From (w/link)", "Status"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=3, column=col, value=h)
    li = ws.cell(row=4, column=2, value="Linkedin")
    li.hyperlink = "https://www.linkedin.com/jobs/view/4012345678"
    ws.cell(row=5, column=2, value="DM Message")            # no link → needs-paste
    rows = jd.read_job_links(ws)
    assert rows[0] == (4, "https://www.linkedin.com/jobs/view/4012345678", "linkedin-guest")
    assert rows[1] == (5, None, "needs-paste")


# G2 — LinkedIn jobId parsing + guest URL
def test_parse_linkedin_job_id():
    jd = _jd()
    assert jd.parse_linkedin_job_id("https://www.linkedin.com/jobs/view/4012345678/") == "4012345678"
    assert jd.parse_linkedin_job_id("https://www.linkedin.com/jobs/search/?currentJobId=987654321&foo=1") == "987654321"
    assert jd.parse_linkedin_job_id("https://example.com/jobs/123") is None  # not LinkedIn's pattern
    assert jd.parse_linkedin_job_id(None) is None


def test_linkedin_guest_url():
    jd = _jd()
    assert jd.linkedin_guest_url("4012345678") == \
        "https://www.linkedin.com/jobs-guest/jobs/api/jobPosting/4012345678"


# G3 — classification into a preflight plan
def test_classify_link_plans(tmp_path):
    jd = _jd()
    assert jd.classify_link("https://www.linkedin.com/jobs/view/4012345678") == "linkedin-guest"
    # LinkedIn URL with no parseable jobId degrades to a browser/paste path, not guest.
    assert jd.classify_link("https://www.linkedin.com/company/acme/jobs/") == "web-fetch"
    assert jd.classify_link("https://greenhouse.io/acme/123") == "web-fetch"
    assert jd.classify_link("") == "needs-paste"
    assert jd.classify_link(None) == "needs-paste"

    # A local file: present → read it; missing → paste.
    present = tmp_path / "jd.pdf"
    present.write_text("a saved jd")
    assert jd.classify_link(str(present)) == "local-file"
    assert jd.classify_link(str(tmp_path / "missing.pdf")) == "needs-paste"


# ── J1: the mandatory end-of-batch summary table (emitted by the ATS script) ─
def _ats():
    return _load_module(
        "ats_score_for_table",
        os.path.join(REPO_ROOT, "templates", "ats_score_template.py"),
    )


def test_summary_why_picks_strongest_strength_and_gap():
    ats = _ats()
    row = ats.summary_row(
        "Acme", "Director of BI", "LEADER",
        87, found=["sql[1]", "business intelligence[3]"],
        missed=["dbt[1]", "snowflake[2]"],
    )
    assert row["why"] == "business intelligence; gap: snowflake"


def test_summary_table_sorts_and_scores_match_tracker():
    ats = _ats()
    rows = [
        ats.summary_row("Low Co", "Analyst", "HANDSON", 71, ["sql[2]"], ["spark[3]"]),
        ats.summary_row("High Co", "Director", "LEADER", 92, ["leadership[3]"], []),
        ats.deferred_row("Gap Co", "Eng", "HANDSON", "JD behind login; pasted not provided"),
    ]
    table = ats.summary_table(rows)
    # The "|---|" separator has no leading space, so it's already excluded here.
    lines = [ln for ln in table.splitlines() if ln.startswith("| ")]
    body = lines[1:]  # skip the header row

    # Sorted high→low, deferred row last with an em dash.
    assert body[0].startswith("| 92 | High Co — Director | LEADER |")
    assert body[1].startswith("| 71 | Low Co — Analyst | HANDSON |")
    assert body[2].startswith("| — | Gap Co — Eng |")
    assert "deferred: JD behind login" in body[2]

    # Each printed score equals the score handed in (the True ATS Score).
    assert "| 92 |" in body[0] and "| 71 |" in body[1]
    # Stats line over scored rows only: min 71, avg round((71+92)/2)=82, max 92.
    assert "MIN 71 · AVG 82 · MAX 92" in table


def test_summary_table_all_deferred_has_no_numeric_stats():
    ats = _ats()
    table = ats.summary_table([ats.deferred_row("X", "Y", None, "no link")])
    assert "MIN — · AVG — · MAX —" in table
