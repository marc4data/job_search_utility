"""Regression tests for Round 5 (v0.4.0): canonical-tracker resolution (T1),
skills-demand index v2 (R1-R4), and the demand review (R5).

Scoring integrity is strengthened, not changed: T1 guarantees the True ATS Score
reaches the canonical tracker and never a user-made backup/OG/copy.
"""
import importlib.util
import os
import sys

import pytest

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TEMPLATES = os.path.join(REPO_ROOT, "templates")
sys.path.insert(0, TEMPLATES)


def _load(name, path):
    spec = importlib.util.spec_from_file_location(name, path)
    m = importlib.util.module_from_spec(spec); spec.loader.exec_module(m)
    return m


def _ats():
    return _load("ats_t1", os.path.join(TEMPLATES, "ats_score_template.py"))


def _make_tracker_dir(tmp_path, *names):
    tdir = tmp_path / "tracker"
    tdir.mkdir()
    for n in names:
        (tdir / n).write_text("x")
    return tmp_path


# ── T1: canonical tracker resolution ─────────────────────────────────────────
def test_t1_resolves_canonical_ignoring_og_backup(tmp_path):
    ats = _ats()
    home = _make_tracker_dir(
        tmp_path,
        "job_search_tracker_marc_alexander.xlsx",
        "job_search_tracker_marc_alexander - OG.xlsx",   # user backup
        "job_search_tracker_marc_alexander_copy.xlsx",   # a copy
        "~$job_search_tracker_marc_alexander.xlsx",       # lock file
    )
    resolved = ats._find_tracker(str(home))
    assert os.path.basename(resolved) == "job_search_tracker_marc_alexander.xlsx"


def test_t1_halts_when_ambiguous(tmp_path):
    ats = _ats()
    home = _make_tracker_dir(
        tmp_path,
        "job_search_tracker_marc_alexander.xlsx",
        "job_search_tracker_jordan_rivera.xlsx",         # two canonical-looking
    )
    with pytest.raises(SystemExit, match="AMBIGUOUS"):
        ats._find_tracker(str(home))


def test_t1_halts_when_none(tmp_path):
    ats = _ats()
    home = _make_tracker_dir(tmp_path, "my notes.xlsx", "tracker - OG.xlsx")
    with pytest.raises(SystemExit, match="Could not resolve"):
        ats._find_tracker(str(home))


def test_t1_backup_name_detection():
    ats = _ats()
    assert ats._is_backup_name("job_search_tracker_x - OG.xlsx")
    assert ats._is_backup_name("job_search_tracker_x_backup.xlsx")
    assert ats._is_backup_name("job_search_tracker_x (1).xlsx")
    assert not ats._is_backup_name("job_search_tracker_marc_alexander.xlsx")


# ── R1-R5: skills-demand index v2 ────────────────────────────────────────────
def _sd():
    return _load("sd_r5", os.path.join(TEMPLATES, "skills_demand.py"))


def _profile(tmp_path):
    p = tmp_path / "profile"; p.mkdir()
    (p / "verified_skills.md").write_text(
        "## Per-company verified experience\n### Cue\n"
        "- **Stack / tools:** dbt, SQL, Tableau\n"
        "- data strategy and data governance leadership\n")
    (p / "bases.py").write_text("BASES = {}\n")
    return str(p)


def test_r1_readable_corpus_filenames(tmp_path):
    sd = _sd()
    home = str(tmp_path)
    sd.record_jd(home, "r15", "Mars Veterinary, Inc.", "Senior Director / People Analytics",
                 "Director", "dbt SQL", date="2026-06-17")
    files = [f for f in os.listdir(os.path.join(home, "docs", "job_descriptions")) if f.endswith(".md")]
    assert files == ["2026-06-17_mars-veterinary-inc_senior-director-people-analytics.md"]


def test_r1_reprocess_updates_not_duplicates(tmp_path):
    sd = _sd()
    home = str(tmp_path)
    sd.record_jd(home, "r15", "Co", "Director", "Director", "dbt", date="2026-06-17")
    sd.record_jd(home, "r15", "Co", "Director", "Director", "dbt SQL", date="2026-06-17")
    mds = [f for f in os.listdir(os.path.join(home, "docs", "job_descriptions")) if f.endswith(".md")]
    assert len(mds) == 1                                   # updated, not duplicated
    job_rows = [r for r in sd.load_index(home) if r["job_id"] == "r15"]
    assert {r["skill"] for r in job_rows} == {"dbt", "sql"}


def test_r2_categories_incl_nontool_from_toolless_jd(tmp_path):
    sd = _sd()
    home = str(tmp_path)
    sd.record_jd(home, "d1", "Co", "Director", "Director",
                 "Own the data strategy and roadmap; mentor the team; executive communication "
                 "to the board; healthcare domain; data governance.", date="2026-06-17")
    cats = {r["category"] for r in sd.load_index(home)}
    assert {"leadership", "technical_competency", "domain"} <= cats
    assert "tool" not in cats                               # no tool tokens in this JD


def test_r3_no_extract_marker_and_reconcile(tmp_path):
    sd = _sd()
    home = str(tmp_path)
    sd.record_jd(home, "x1", "Co", "Analyst", "IC", "We value curiosity and grit.", date="2026-06-17")
    rows = sd.load_index(home)
    assert [r["skill"] for r in rows] == [sd.NO_EXTRACT]    # one explicit marker, not silence
    assert sd.reconcile(rows) == {"processed": 1, "no_extract": 1}


def test_r4_stoplist_strips_claude(tmp_path):
    sd = _sd()
    home = str(tmp_path)
    sd.record_jd(home, "n1", "Co", "Engineer", "IC", "Use Claude and dbt and SQL.", date="2026-06-17")
    skills = {r["skill"] for r in sd.load_index(home)}
    assert "claude" not in skills and {"dbt", "sql"} <= skills


def test_r5_demand_workbook_has_coverage(tmp_path):
    import openpyxl
    sd = _sd()
    home = str(tmp_path)
    prof = _profile(tmp_path)
    sd.record_jd(home, "j1", "Co", "Director", "Director",
                 "dbt SQL Snowflake; data strategy; healthcare", date="2026-06-17")
    out = sd.write_demand_workbook(home, prof)
    ws = openpyxl.load_workbook(out)["Skills Demand"]
    headers = [c.value for c in ws[1]]
    assert headers == ["Category", "Skill", "Jobs", "% of jobs", "Coverage", "Levels"]
    body = [[c.value for c in row] for row in ws.iter_rows(min_row=2)]
    cats = {r[0] for r in body}
    assert "tool" in cats and ("leadership" in cats or "domain" in cats)   # non-tool surfaced
    covs = {r[4] for r in body}
    assert covs <= {"Covered", "Weak", "Gap"}
    # dbt is backed in the profile → Covered; Snowflake absent → Gap
    by_skill = {r[1]: r[4] for r in body}
    assert by_skill.get("dbt") == "Covered"
    assert by_skill.get("Snowflake") == "Gap"
