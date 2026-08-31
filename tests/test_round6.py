"""Regression tests for Round 6 (v0.5.0): the manual-JD fallback + archive (W1/W3),
the no-prompt retrieval contract (W2), and the cover-letter presumption rule (W4).

The deterministic surface is what's locked here: filename parsing, company/title
matching (including the near-misses that must NOT match), reading each supported
format, the archive sweep's safety rule, and the shipped prompt contracts. The
model actually fetching a JD stays a clean-room concern.
"""
import json
import os
import sys

import pytest

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TEMPLATES = os.path.join(REPO_ROOT, "templates")
sys.path.insert(0, TEMPLATES)

from conftest import _load_module

SKILL_PROCESS = os.path.join(REPO_ROOT, "skills", "process-opportunities", "SKILL.md")
SKILL_SETUP = os.path.join(REPO_ROOT, "skills", "setup-profile", "SKILL.md")
RULES = os.path.join(REPO_ROOT, "references", "process_rules.md")


def _mj():
    return _load_module("manual_jd_under_test", os.path.join(TEMPLATES, "manual_jd.py"))


def _jd():
    return _load_module("jd_retrieval_w1", os.path.join(TEMPLATES, "jd_retrieval.py"))


def _drop(tmp_path, name, text="Job description body."):
    """Write a manual JD into <home>/docs/manual_job_descriptions/."""
    d = tmp_path / "docs" / "manual_job_descriptions"
    d.mkdir(parents=True, exist_ok=True)
    path = d / name
    if name.lower().endswith(".docx"):
        from docx import Document
        doc = Document()
        doc.add_paragraph(text)
        doc.save(str(path))
    else:
        path.write_text(text, encoding="utf-8")
    return path


# ── W1: filename parsing ────────────────────────────────────────────────────
def test_parse_filename_strategy_and_optional_date():
    mj = _mj()
    p = mj.parse_manual_filename("20260830 Acme Health - Director of Analytics.docx")
    assert (p["date"], p["company"], p["title"], p["ext"]) == \
        ("2026-08-30", "Acme Health", "Director of Analytics", ".docx")

    # Dashed/underscored dates and no date at all are all accepted.
    assert mj.parse_manual_filename("2026-08-30 Acme - BI Lead.md")["date"] == "2026-08-30"
    assert mj.parse_manual_filename("Acme - BI Lead.txt")["date"] is None

    # The FIRST separator splits, so a dash inside the title survives.
    p2 = mj.parse_manual_filename("20260830 Acme - Director - Analytics.docx")
    assert (p2["company"], p2["title"]) == ("Acme", "Director - Analytics")

    # No company/title separator → not following the strategy → never matched.
    assert mj.parse_manual_filename("some notes.docx") is None


# ── W1: matching, including the near-misses that must NOT match ─────────────
def test_match_is_forgiving_on_company_and_strict_across_roles(tmp_path):
    mj = _mj()
    home = str(tmp_path)
    _drop(tmp_path, "20260830 Acme Health Inc - Director of Analytics.docx")

    # Legal suffixes and title connector words don't have to match exactly.
    assert mj.find_manual_jd(home, "Acme Health", "Director, Analytics")
    assert mj.find_manual_jd(home, "Acme Health, Inc.", "Director of Analytics")

    # A different role at the same company must NOT borrow this file.
    assert mj.find_manual_jd(home, "Acme Health", "Marketing Manager") is None
    # Nor the same title at a different company.
    assert mj.find_manual_jd(home, "Globex", "Director of Analytics") is None


def test_match_picks_the_right_file_among_siblings(tmp_path):
    mj = _mj()
    home = str(tmp_path)
    _drop(tmp_path, "20260801 Acme - Data Engineer.docx")
    _drop(tmp_path, "20260830 Acme - Director of Analytics.docx")
    hit = mj.find_manual_jd(home, "Acme", "Director of Analytics")
    assert hit["filename"] == "20260830 Acme - Director of Analytics.docx"


# ── W1: reading each supported format ───────────────────────────────────────
@pytest.mark.parametrize("name", [
    "20260830 Acme - Director of Analytics.docx",
    "20260830 Acme - Director of Analytics.md",
    "20260830 Acme - Director of Analytics.txt",
])
def test_manual_fallback_reads_supported_formats(tmp_path, name):
    mj = _mj()
    _drop(tmp_path, name, text="Must have dbt and Snowflake.")
    got = mj.manual_fallback(str(tmp_path), "Acme", "Director of Analytics")
    assert "dbt and Snowflake" in got["text"]


def test_unreadable_match_is_reported_not_guessed(tmp_path):
    mj = _mj()
    _drop(tmp_path, "20260830 Acme - Director of Analytics.pdf", text="%PDF-1.4")
    got = mj.manual_fallback(str(tmp_path), "Acme", "Director of Analytics")
    # Matched, but no text and a reason naming the file — never a silent miss.
    assert got["text"] is None
    assert got["filename"] in got["reason"]
    with pytest.raises(ValueError):
        mj.read_manual_jd(got["path"])


def test_no_folder_and_no_match_are_both_just_none(tmp_path):
    mj = _mj()
    assert mj.list_manual_jds(str(tmp_path)) == []
    assert mj.manual_fallback(str(tmp_path), "Acme", "Director") is None


def test_office_lock_files_are_ignored(tmp_path):
    mj = _mj()
    _drop(tmp_path, "~$20260830 Acme - Director of Analytics.docx", text="lock")
    assert mj.list_manual_jds(str(tmp_path)) == []


# ── W1: the preflight plan sees the fallback ────────────────────────────────
def test_plan_for_role_upgrades_needs_paste_to_manual_file(tmp_path):
    jd, mj = _jd(), _mj()
    home = str(tmp_path)
    _drop(tmp_path, "20260830 Acme - Director of Analytics.docx")

    # No link at all, but a manual file is waiting → manual-file, not needs-paste.
    plan, match = jd.plan_for_role(None, home=home, company="Acme", role="Director of Analytics")
    assert plan == "manual-file" and match["filename"].endswith(".docx")

    # A role with a good link keeps its fetch plan but still carries the fallback,
    # so a failed fetch never has to stop and ask.
    plan, match = jd.plan_for_role("https://acme.example/jobs/1", home=home,
                                   company="Acme", role="Director of Analytics")
    assert plan == "web-fetch" and match is not None

    # Without a home there is no fallback to find — behavior is unchanged.
    assert jd.plan_for_role(None) == ("needs-paste", None)


def test_read_job_rows_attaches_company_role_and_manual(tmp_path):
    import openpyxl

    jd = _jd()
    home = str(tmp_path)
    _drop(tmp_path, "20260830 Acme - Director of Analytics.docx")
    wb = openpyxl.Workbook()
    ws = wb.active
    for col, h in enumerate(["Company", "Role / Title", "Sourced From (w/link)"], start=1):
        ws.cell(row=3, column=col, value=h)
    ws.cell(row=4, column=1, value="Acme")
    ws.cell(row=4, column=2, value="Director of Analytics")
    ws.cell(row=4, column=3, value="DM Message")          # no resolvable link

    rows = jd.read_job_rows(ws, home=home)
    assert rows[0]["company"] == "Acme" and rows[0]["role"] == "Director of Analytics"
    assert rows[0]["plan"] == "manual-file"
    assert rows[0]["manual"]["path"].endswith("20260830 Acme - Director of Analytics.docx")


# ── W3: archiving used manual JDs ───────────────────────────────────────────
def test_archive_moves_used_files_without_overwriting(tmp_path):
    mj = _mj()
    home = str(tmp_path)
    used = _drop(tmp_path, "20260830 Acme - Director of Analytics.docx")
    moved, left = mj.archive_manual_jds(home, [str(used)])
    assert not left and not used.exists()
    assert os.path.basename(moved[0]) == "20260830 Acme - Director of Analytics.docx"
    # The archive is not part of the live drop folder any more.
    assert mj.list_manual_jds(home) == []

    # A same-named file archived later is suffixed, never overwritten.
    again = _drop(tmp_path, "20260830 Acme - Director of Analytics.docx", text="v2")
    moved2, _ = mj.archive_manual_jds(home, [str(again)])
    assert moved2[0].endswith("(2).docx") and os.path.exists(moved[0])


def test_stale_sweep_archives_previous_runs_but_spares_fresh_drops(tmp_path):
    mj = _mj()
    home = str(tmp_path)
    old = _drop(tmp_path, "20260801 Acme - Data Engineer.docx")       # a prior run used this
    fresh = _drop(tmp_path, "20260830 Globex - BI Lead.docx")         # dropped in for today

    index = [{"company": "Acme", "role": "Data Engineer"}]            # skills-demand index
    stale = mj.stale_manual_jds(home, active_paths=[], processed=index)
    assert [p["filename"] for p in stale] == [old.name]

    # A file this batch is about to use is never swept, even if it's in the index.
    assert mj.stale_manual_jds(home, active_paths=[str(old)], processed=index) == []
    assert fresh.exists()


# ── W2: the no-prompt retrieval contract ────────────────────────────────────
def test_workspace_settings_template_preapproves_web_access():
    with open(os.path.join(TEMPLATES, "workspace_settings_template.json")) as f:
        cfg = json.load(f)
    assert {"WebFetch", "WebSearch"} <= set(cfg["permissions"]["allow"])


def test_retrieval_contract_forbids_mid_run_prompts():
    text = open(SKILL_PROCESS).read() + open(RULES).read()
    assert "workspace_settings_template.json" in text
    assert "Never prompt during retrieval" in text
    # The fallback must be reached for before any paste request.
    assert "manual_fallback" in text and "manual_job_descriptions" in text


def test_setup_creates_the_manual_folder_and_settings():
    text = open(SKILL_SETUP).read()
    assert "docs/manual_job_descriptions" in text
    assert "workspace_settings_template.json" in text
    assert "manual_jd_README_template.md" in text


# ── W4: the cover-letter presumption rule ───────────────────────────────────
def test_cover_letter_rule_bans_prescribing_to_the_company():
    rules = open(RULES).read()
    assert "presumption rule" in rules.lower()
    assert "Never assert what you would do at the company" in rules
    # The old strategy — diagnose their problem, then propose what you'd do — is gone.
    assert "their problem, not your bio" not in rules
    assert "what you'd prioritize in the role" not in rules
    # And the new one is stated: enthusiasm + résumé/JD match.
    assert "Para 2 — the match" in rules.lower() or "**Para 2 — the match.**" in rules


def test_shipped_cover_letter_placeholders_follow_the_new_strategy():
    text = open(os.path.join(TEMPLATES, "build_batch_template.py")).read()
    assert "THE MATCH" in text and "THE CLOSE" in text
    assert "what you'd prioritize" not in text
    assert "ONE vivid story that maps to the problem" not in text
