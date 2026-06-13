"""A1 regression: the canonical workspace layout + new path references work.

Builds the §2 tree in a temp dir, drops the engine into .system/engine/, copies
the two templates into .system/scripts/<date>/ (as a real run would), executes
them, and asserts:
  - built .docx land in docs/current/ (not the working-folder root),
  - the ATS script finds the tracker under tracker/ and writes Resume Score,
  - HOME is resolved via the .system/ ancestor walk (deep script nesting is fine).
"""
import os
import shutil
import subprocess
import sys

import openpyxl

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

DUMMY_PROFILE = '''\
PROFILE = {
    "name": "Sam Sample",
    "contact": "Somewhere, ST  \\u2022  (555) 123-4567  \\u2022  sam@example.invalid",
    "cl_contact": "Somewhere, ST  |  (555) 123-4567  |  sam@example.invalid",
    "links": [("LinkedIn", "https://linkedin.invalid/in/sam")],
    "education": ["BS in Data Analytics  \\u2022  State University"],
    "certifications": ["Certifications: Tableau Desktop Specialist"],
    "additional_experience": ["Data Analyst  \\u2022  First Job Co"],
}
'''

DUMMY_BASES = '''\
BASES = {
    "LEADER": {
        "career_highlights": ["Built the analytics function and delivered through 3x growth."],
        "experience": [
            {"role": "Director of Analytics", "company_loc": "Sample Corp  \\u2022  Somewhere, ST",
             "dates": "2021 - Present", "desc": "Leads the BI org.",
             "bullets": ["Set analytics strategy and the company-wide KPI framework."]},
        ],
        "tech_expertise": [("Analytic Tools", "Tableau, Power BI, SQL, Python")],
        "projects": None,
        "certs_before_edu": False,
    },
}
'''


def _make_tracker(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Applications"
    ws.cell(row=1, column=1, value="Job Application Tracker")
    for col, head in enumerate(["#", "Company", "Role / Title", "Resume Score", "Status"], start=1):
        ws.cell(row=3, column=col, value=head)
    ws.cell(row=4, column=2, value="Example Co")
    ws.cell(row=4, column=3, value="Director of Analytics")
    wb.save(str(path))


def test_new_workspace_layout_smoke_build(tmp_path):
    home = tmp_path / "job_search"
    # the canonical tree
    (home / ".system" / "engine").mkdir(parents=True)
    date_dir = home / ".system" / "scripts" / "2026-06-13"
    date_dir.mkdir(parents=True)
    (home / "profile" / "history").mkdir(parents=True)
    (home / "tracker").mkdir()
    (home / "docs" / "current").mkdir(parents=True)
    (home / "docs" / "submitted").mkdir(parents=True)

    # engine + profile + tracker
    shutil.copy(os.path.join(REPO, "engine", "build_docs.py"), home / ".system" / "engine" / "build_docs.py")
    (home / "profile" / "profile.py").write_text(DUMMY_PROFILE)
    (home / "profile" / "bases.py").write_text(DUMMY_BASES)
    _make_tracker(home / "tracker" / "job_search_tracker_sam_sample.xlsx")

    # the two templates, copied in as a real run would (deep under .system/scripts/<date>/)
    shutil.copy(os.path.join(REPO, "templates", "build_batch_template.py"), date_dir / "build_batch_2026-06-13.py")
    shutil.copy(os.path.join(REPO, "templates", "ats_score_template.py"), date_dir / "ats_score_2026-06-13.py")

    # run the build (executes do_role at module level)
    r1 = subprocess.run([sys.executable, str(date_dir / "build_batch_2026-06-13.py")],
                        capture_output=True, text=True)
    assert r1.returncode == 0, f"build failed:\n{r1.stdout}\n{r1.stderr}"

    current = home / "docs" / "current"
    built = sorted(p.name for p in current.glob("*.docx"))
    assert any("Resume" in n for n in built), f"no résumé in docs/current/: {built}"
    assert any("Cover Letter" in n for n in built), f"no cover letter in docs/current/: {built}"
    # nothing leaked to the working-folder root (old flat behavior)
    assert not list(home.glob("*.docx")), "documents leaked to the working-folder root"

    # run the scorer (writes Resume Score into tracker/)
    r2 = subprocess.run([sys.executable, str(date_dir / "ats_score_2026-06-13.py")],
                        capture_output=True, text=True)
    assert r2.returncode == 0, f"scorer failed:\n{r2.stdout}\n{r2.stderr}"

    wb = openpyxl.load_workbook(home / "tracker" / "job_search_tracker_sam_sample.xlsx")
    ws = wb["Applications"]
    score_col = [c.value for c in ws[3]].index("Resume Score") + 1
    score = ws.cell(row=4, column=score_col).value
    assert isinstance(score, int) and 0 < score < 100, f"expected an honest score, got {score!r}"
