"""make_profile_workbook.py — DEV tool (not shipped in the .plugin).

Builds templates/profile_workbook_template.xlsx from the repo's *.example.* profile
files, so the shipped template is the example profile in friendly Excel form and the
workbook↔files round-trip is proven against real structure.

Run:  python3 tools/make_profile_workbook.py
"""
import importlib.util
import os
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(REPO, "templates", "profile_workbook_template.xlsx")
NAVY, GREY = "FF1F3864", "FFF2F2F2"
HFONT = Font(name="Arial", size=10, bold=True, color="FFFFFFFF")
WRAP = Alignment(wrap_text=True, vertical="top")


def _load(name, path):
    spec = importlib.util.spec_from_file_location(name, path)
    m = importlib.util.module_from_spec(spec); spec.loader.exec_module(m)
    return m


PROFILE = _load("ex_profile", os.path.join(REPO, "profile", "profile.example.py")).PROFILE
BASES = _load("ex_bases", os.path.join(REPO, "profile", "bases.example.py")).BASES


def sheet(wb, title, headers, widths):
    ws = wb.create_sheet(title)
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = HFONT; c.fill = PatternFill("solid", fgColor=NAVY)
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    return ws


def put(ws, r, values):
    for c, v in enumerate(values, start=1):
        cell = ws.cell(row=r, column=c, value=v); cell.alignment = WRAP


wb = openpyxl.Workbook()

# How to use
ws = wb.active; ws.title = "How to use"; ws.column_dimensions["A"].width = 100
how = [
    "YOUR PROFILE WORKBOOK — the one file you edit to tune your résumés over time.",
    "",
    "You never edit the Python/Markdown files; Claude compiles those from this workbook.",
    "Edit a tab, then say \"update my profile\" — Claude recompiles profile.py, bases.py,",
    "and verified_skills.md, and runs the truthfulness check.",
    "",
    "Tabs:",
    "  Identity         → your name/contact/links/education/certs (profile.py)",
    "  Skills Matrix    → which TOOL you used at which COMPANY (the truth source)",
    "  Experience       → per-company domain / what you did / leadership",
    "  Career Highlights, Roles, Role Bullets, Tech Expertise, Projects, Bases Config",
    "                   → your two base résumés (LEADER and HANDSON)",
    "",
    "Skills Matrix flags: Portfolio-only? = used in a project, never an employer.",
    "                     Never-used?    = must never appear on a résumé.",
]
for i, line in enumerate(how, start=1):
    ws.cell(row=i, column=1, value=line).font = Font(name="Arial", size=10, bold=line.endswith(":"))

# Identity
ws = sheet(wb, "Identity", ["Field", "Value"], [24, 92])
loc, phone, email = (PROFILE["contact"].split("•") + ["", "", ""])[:3]
ident = [("Name", PROFILE["name"]), ("Location", loc.strip()), ("Phone", phone.strip()),
         ("Email", email.strip())]
for label, url in PROFILE["links"]:
    ident.append((f"Link: {label}", url))
for i, e in enumerate(PROFILE["education"], 1):
    ident.append((f"Education {i}", e))
for line in PROFILE["certifications"]:
    if line.lower().startswith("training:"):
        ident.append(("Training", line.split(":", 1)[1].strip()))
    elif line.lower().startswith("certifications:"):
        ident.append(("Certifications", line.split(":", 1)[1].strip()))
for i, a in enumerate(PROFILE.get("additional_experience", []), 1):
    ident.append((f"Additional Experience {i}", a))
for r, (k, v) in enumerate(ident, start=2):
    put(ws, r, [k, v]); ws.cell(row=r, column=1).font = Font(name="Arial", size=10, bold=True)

# Skills Matrix (coherent example consistent with the example verified_skills)
ws = sheet(wb, "Skills Matrix",
           ["Skill / Tool", "Company", "Depth", "Portfolio-only?", "Never-used?", "Notes"],
           [20, 24, 12, 16, 14, 40])
matrix = [
    ("SQL", "Company A", "Expert", "", "", ""),
    ("Python", "Company A", "Working", "", "", ""),
    ("dbt", "Company A", "Working", "", "", ""),
    ("BigQuery", "Company A", "Working", "", "", ""),
    ("Tableau", "Company B", "Expert", "", "", ""),
    ("Power BI", "Company B", "Working", "", "", ""),
    ("Snowflake", "", "Learning", "Yes", "", "Portfolio project only — not at an employer"),
    ("Looker", "", "", "", "", "On skills list — confirm which employer"),
]
for r, row in enumerate(matrix, start=2):
    put(ws, r, list(row))
dv_depth = DataValidation(type="list", formula1='"Expert,Working,Learning"', allowBlank=True)
dv_yn = DataValidation(type="list", formula1='"Yes,No"', allowBlank=True)
ws.add_data_validation(dv_depth); ws.add_data_validation(dv_yn)
dv_depth.add("C2:C500"); dv_yn.add("D2:D500"); dv_yn.add("E2:E500")

# Experience (per-company meta for verified_skills.md)
ws = sheet(wb, "Experience", ["Company", "Title", "Dates", "Domain", "What you did", "Leadership"],
           [18, 26, 16, 22, 36, 30])
put(ws, 2, ["Company A", "Director of Analytics", "2022 - Present",
            "e.g. fintech / e-commerce", "Built the analytics function from scratch", "People management"])
put(ws, 3, ["Company B", "Analytics Manager", "2018 - 2022",
            "e.g. healthcare", "Overhauled reporting and governance", "Player-coach"])

# Bases tabs (from bases.example)
ch = sheet(wb, "Career Highlights", ["Base", "Highlight"], [12, 100]); rch = 2
roles = sheet(wb, "Roles", ["Base", "Company • Loc", "Dates", "Role", "Desc"], [12, 30, 18, 28, 50]); rr = 2
rb = sheet(wb, "Role Bullets", ["Base", "Company • Loc", "Bullet"], [12, 30, 90]); rrb = 2
te = sheet(wb, "Tech Expertise", ["Base", "Label", "Items"], [12, 24, 80]); rte = 2
pj = sheet(wb, "Projects", ["Base", "Project", "Tech", "Bullet"], [12, 30, 40, 70]); rpj = 2
cfg = sheet(wb, "Bases Config", ["Base", "certs_before_edu"], [12, 18]); rcfg = 2

for base, block in BASES.items():
    for h in block["career_highlights"]:
        put(ch, rch, [base, h]); rch += 1
    for e in block["experience"]:
        put(roles, rr, [base, e["company_loc"], e["dates"], e["role"], e["desc"]]); rr += 1
        for b in e["bullets"]:
            put(rb, rrb, [base, e["company_loc"], b]); rrb += 1
    for label, items in block["tech_expertise"]:
        put(te, rte, [base, label, items]); rte += 1
    for p in (block.get("projects") or []):
        for b in p["bullets"]:
            put(pj, rpj, [base, p["name"], p["tech"], b]); rpj += 1
    put(cfg, rcfg, [base, "Yes" if block.get("certs_before_edu") else "No"]); rcfg += 1

wb.save(OUT)
print(f"Built {OUT}")
print(f"Tabs: {wb.sheetnames}")
